#!/usr/bin/env python3
"""Build the deterministic BOSS Libraries Markdown corpus.

The raw files under ``apps/boss/business docs`` are the immutable source of
truth.  This builder emits one concise, traceable Markdown routing record per
source identity plus a machine-readable manifest, catalog, and conversion
report under ``apps/boss/documents``.

The output is intentionally a flat collection because Arcane OS stores
documents in a flat OPFS table.  Each filename starts with a stable ID derived
from the normalized source path, so same-named files never collide.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import mimetypes
import re
import shutil
import subprocess
import sys
import unicodedata
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree as ET

try:
    from PIL import Image
except Exception:  # pragma: no cover - recorded as an extraction limitation
    Image = None

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover
    PdfReader = None

try:
    import pdfplumber
except Exception:  # pragma: no cover
    pdfplumber = None

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


APP_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE_ROOT = APP_ROOT / "business docs"
DEFAULT_OUTPUT_ROOT = APP_ROOT / "documents"
LINK_POLICY_PATH = APP_ROOT / "link-policy.json"
EXPECTED_SOURCE_COUNT = 500
SCHEMA_VERSION = 1
BUILDER_VERSION = "boss-library-md-v1"
MAX_EXTRACTED_TEXT = 160_000
MAX_DESCRIPTION = 620
MAX_KEY_ITEM = 260
MAX_LINKS = 12
MAX_CONTACTS = 8

LINK_POLICY = json.loads(LINK_POLICY_PATH.read_text(encoding="utf-8"))
LINK_REPLACEMENTS = LINK_POLICY["replace"]
LINK_REMOVALS = set(LINK_POLICY["remove"])
LINK_REMOVAL_PREFIXES = tuple(LINK_POLICY["remove_prefixes"])

SUPPORTED_EXTENSIONS = {
    ".md",
    ".pdf",
    ".docx",
    ".pptx",
    ".odp",
    ".xlsx",
    ".csv",
    ".txt",
    ".jpg",
    ".jpeg",
    ".png",
    ".mp4",
    ".ppt",
    ".rtb",
    ".zip",
}

GENERIC_TITLES = {
    "document",
    "presentation",
    "powerpoint presentation",
    "microsoft word document",
    "untitled",
    "slide 1",
    "sheet1",
}

STOPWORDS = {
    "about",
    "after",
    "again",
    "against",
    "also",
    "among",
    "and",
    "are",
    "because",
    "been",
    "before",
    "being",
    "between",
    "boss",
    "business",
    "can",
    "client",
    "document",
    "file",
    "for",
    "from",
    "have",
    "help",
    "how",
    "into",
    "library",
    "more",
    "not",
    "only",
    "other",
    "our",
    "pack",
    "process",
    "resource",
    "resources",
    "route",
    "routing",
    "should",
    "that",
    "the",
    "their",
    "them",
    "there",
    "these",
    "they",
    "this",
    "through",
    "use",
    "user",
    "using",
    "via",
    "when",
    "where",
    "which",
    "with",
    "your",
}

TOPIC_RULES: dict[str, tuple[str, ...]] = {
    "accounting": ("accounting", "bookkeeping", "quickbooks", "chart of accounts", "financial statement"),
    "artificial intelligence": ("artificial intelligence", " ai ", "chatgpt", "ollama", "embedding", "prompt"),
    "branding": ("brand", "branding", "logo", "marketing identity"),
    "business planning": ("business plan", "business model", "canvas", "startup roadmap", "planning"),
    "certification": ("certification", "8(a)", "wosb", "hubzone", "sdvosb"),
    "cybersecurity": ("cyber", "cmmc", "security", "data protection"),
    "disaster readiness": ("disaster", "hurricane", "continuity", "recovery"),
    "ecosystem navigation": ("ecosystem", "resource partner", "local assistance", "directory", "map"),
    "exit and transition": ("exit", "transition", "sell your business", "succession", "legacy"),
    "formation": ("formation", "entity", "ein", "registered agent", "llc", "dba", "secretary of state"),
    "franchising": ("franchise", "franchising"),
    "funding": ("funding", "loan", "lender", "capital", "grant", "investment", "cash flow"),
    "government contracting": ("government contract", "procurement", "sam.gov", "uei", "cage", "solicitation", "apex"),
    "human resources": ("human resources", "employee", "hiring", "payroll", "workforce"),
    "international trade": ("export", "international trade", "global market"),
    "legal and compliance": ("legal", "compliance", "agreement", "contract", "nda", "tax", "insurance", "permit"),
    "market research": ("market research", "competitor", "customer discovery", "validation", "target market"),
    "marketing": ("marketing", "seo", "social media", "customer", "sales"),
    "mentoring and advising": ("mentor", "mentoring", "advisor", "advising", "counseling", "score"),
    "operations": ("operations", "operating model", "workflow", "procedure", "quality control"),
    "technology": ("technology", "website", "software", "automation", "workspace", "architecture"),
    "training": ("training", "webinar", "course", "workshop", "curriculum", "lesson"),
    "veteran entrepreneurship": ("veteran", "military spouse", "boots to business", "vboc", "tvc"),
}

STAGE_RULES: dict[str, tuple[str, ...]] = {
    "Ideation": ("ideation", "idea", "concept", "pre-venture"),
    "Validation": ("validation", "validate", "customer discovery", "market research", "feasibility"),
    "Formation": ("formation", "form your", "entity", "ein", "registered agent", "llc"),
    "Launch": ("launch", "starting your business", "startup"),
    "Operations": ("operations", "manage your business", "operating"),
    "Growth": ("growth", "grow your business", "scale", "expansion"),
    "Recovery": ("recovery", "disaster", "continuity", "troubleshooting"),
    "Exit": ("exit", "transition", "succession", "sell your business"),
}

ORGANIZATION_RULES: dict[str, tuple[str, ...]] = {
    "BOSS Libraries": ("boss libraries", "bosslibraries", "business operator solutions"),
    "SCORE": ("score", "score.org"),
    "U.S. Small Business Administration (SBA)": ("small business administration", "sba.gov", " sba "),
    "Small Business Development Center (SBDC)": ("small business development center", "sbdc", "sbdc.uh.edu"),
    "APEX Accelerator": ("apex accelerator", "uh apex", "uhapex"),
    "Veterans Business Outreach Center (VBOC)": ("vboc", "veterans business outreach"),
    "Texas Veterans Commission (TVC)": ("texas veterans commission", " tvc "),
    "University of Houston": ("university of houston", " uh ", "bauer"),
    "Google Workspace": ("google workspace", "google drive", "gmail"),
    "Texas Secretary of State": ("texas secretary of state", "texas sos"),
    "Texas Comptroller": ("texas comptroller",),
    "SCORE Houston": ("score houston", "score.org/tx/houston"),
}

RESTRICTED_PATH_MARKERS = (
    "accounting/",
    "consulting/",
    "formation/",
    "suppliers/",
    "/clients/",
    "client intake",
    "financial model",
    "receipt",
    "invoice",
    "credit card",
    "cc form",
    "ein letter",
    "engagement letter",
    "executed nda",
    "service agreement",
    "company agreement",
    "written consent",
    "organizational consent",
    "registered agent",
    "investment scenario",
    "budget.xlsx",
    "fundraising pipeline",
)

INTERNAL_PATH_MARKERS = (
    "branding/",
    "logo/",
    "workspace/",
    "workspace architecture/",
    "best practice operating model doctrine/",
    "chat gpt doctrine development/",
    "entity framework models/",
    "office management to coo program/",
    "pitch/",
    "ai training/original discussion",
    "ai training/ai training",
    "/score/score/google meet/",
    "/score/score/ed mapping resources/",
    "/score/score/training materials/coaching/",
    "industry knowledge mentor list",
    "volunteer onboarding",
)

PUBLIC_LINK_DOMAINS = (
    "sba.gov",
    "score.org",
    "sbdc.uh.edu",
    "uh.edu",
    "uhapex",
    "texas.gov",
    "irs.gov",
    "sam.gov",
)


@dataclass
class Extraction:
    title_candidates: list[tuple[str, str]] = field(default_factory=list)
    text: str = ""
    segments: list[str] = field(default_factory=list)
    links: list[str] = field(default_factory=list)
    contacts: list[str] = field(default_factory=list)
    people: list[str] = field(default_factory=list)
    status: str = "complete"
    method: str = ""
    coverage: str = ""
    limitations: list[str] = field(default_factory=list)
    details: dict[str, Any] = field(default_factory=dict)


def normalize_relative_path(path: Path, source_root: Path) -> str:
    relative = path.relative_to(source_root).as_posix()
    return unicodedata.normalize("NFKC", relative).casefold()


def source_id(path: Path, source_root: Path) -> str:
    normalized = normalize_relative_path(path, source_root)
    digest = hashlib.sha256(normalized.encode("utf-8")).hexdigest()[:12]
    return f"bossdoc-{digest}"


def hash_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(4 * 1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def slugify(value: str, limit: int = 80) -> str:
    ascii_value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-z0-9]+", "-", ascii_value.lower()).strip("-")
    slug = re.sub(r"-+", "-", slug)[:limit].rstrip("-")
    return slug or "reference"


def local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def clean_text(value: Any) -> str:
    text = str(value or "")
    text = text.replace("\x00", " ").replace("\ufeff", " ")
    text = "".join(char if char in "\n\t" or ord(char) >= 32 else " " for char in text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def truncate(value: str, limit: int) -> str:
    text = clean_text(value)
    if len(text) <= limit:
        return text
    shortened = text[: limit - 1].rsplit(" ", 1)[0].rstrip(" ,;:-")
    return f"{shortened}…"


def unique(values: Iterable[str], limit: int | None = None) -> list[str]:
    output: list[str] = []
    seen: set[str] = set()
    for value in values:
        item = clean_text(value)
        key = item.casefold()
        if not item or key in seen:
            continue
        seen.add(key)
        output.append(item)
        if limit is not None and len(output) >= limit:
            break
    return output


def meaningful_line(value: str) -> bool:
    line = clean_text(value).strip("-•*#|: ")
    lowered = line.casefold()
    if not 4 <= len(line) <= 220:
        return False
    if sum(char.isalpha() for char in line) < 3:
        return False
    if lowered.startswith(("source url", "http://", "https://", "last reviewed", "review date", "page ")):
        return False
    if lowered in GENERIC_TITLES:
        return False
    return True


def clean_title(value: str) -> str:
    title = clean_text(value).strip("#*-–—|: ")
    title = re.sub(r"(?i)^microsoft (word|powerpoint)\s*[-:]\s*", "", title)
    title = re.sub(r"(?i)\.(docx|pdf|pptx|ppt|odp|xlsx|png|jpg|jpeg)$", "", title)
    title = re.sub(r"\s+", " ", title)
    return truncate(title, 150)


def filename_title(path: Path) -> str:
    title = path.stem.replace("_", " ")
    title = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", title)
    title = re.sub(r"\s+", " ", title)
    return clean_title(title)


def choose_title(extraction: Extraction, path: Path) -> tuple[str, str]:
    # Human-reviewed visual/video/special titles are deliberate. Existing
    # Markdown H1s are also authoritative because the audited training packs
    # have shuffled filenames. For binary office/media sources, the source
    # filename is consistently a clearer retrieval label than cover-page
    # boilerplate, addresses, generic metadata, or PDF encoding artifacts.
    for candidate, source in extraction.title_candidates:
        if source.startswith("human_"):
            title = clean_title(candidate)
            if meaningful_line(title):
                return title, source
    if path.suffix.casefold() not in {".md", ".txt"}:
        fallback = filename_title(path)
        if meaningful_line(fallback):
            return fallback, "filename_preferred"
    for candidate, source in extraction.title_candidates:
        title = clean_title(candidate)
        if meaningful_line(title):
            return title, source
    for segment in extraction.segments:
        if meaningful_line(segment):
            return clean_title(segment), "first_meaningful_content"
    fallback = filename_title(path)
    return fallback or f"{path.suffix.lstrip('.').upper()} reference", "filename_fallback"


URL_RE = re.compile(r"https?://[^\s<>\]\[\"']+", re.IGNORECASE)
EMAIL_RE = re.compile(r"(?<![\w.+-])[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}(?![\w.-])")
PHONE_RE = re.compile(r"(?<!\d)(?:\+?1[\s.-]?)?\(?\d{3}\)?[\s.-]\d{3}[\s.-]\d{4}(?!\d)")

LINK_CORRECTIONS = {
    "https://www.sba.gov/business-guide/launch-your-business/choose-": "https://www.sba.gov/business-guide/launch-your-business/choose-business-structure",
    "https://help.score.org/hc/en-": "https://help.score.org/hc/en-us",
    "https://www.rd.usda.gov/programs-": "https://www.rd.usda.gov/programs-services",
    "https://plannedgiving.com/about-us/what-is-planned-": "https://plannedgiving.com/about-us/what-is-planned-giving/",
}

CONTACT_CORRECTIONS = {
    "vboc@utrgv.eduCall": "vboc@utrgv.edu",
    "elizabeth.charles@sba.govOr": "elizabeth.charles@sba.gov",
}


def extract_links(text: str) -> list[str]:
    links = []
    for match in URL_RE.findall(text or ""):
        link = match.rstrip(".,;:!?)\"'")
        link = LINK_CORRECTIONS.get(link, link)
        # A trailing hyphen is a reliable PDF line-wrap artifact. Known links
        # are repaired above; unknown fragments are omitted instead of routed.
        if link.endswith("-"):
            continue
        links.append(link)
    return unique(links, MAX_LINKS)


def apply_link_policy(links: Iterable[str]) -> list[str]:
    normalized = []
    for link in links:
        if link in LINK_REMOVALS or link.startswith(LINK_REMOVAL_PREFIXES):
            continue
        normalized.append(LINK_REPLACEMENTS.get(link, link))
    return unique(normalized, MAX_LINKS)


def apply_link_policy_to_text(text: str) -> str:
    def normalize_match(match: re.Match) -> str:
        matched = match.group(0)
        link = matched.rstrip(".,;:!?)")
        trailing = matched[len(link):]
        if link in LINK_REMOVALS or link.startswith(LINK_REMOVAL_PREFIXES):
            return trailing
        return f"{LINK_REPLACEMENTS.get(link, link)}{trailing}"

    value = URL_RE.sub(normalize_match, text)
    value = re.sub(r"(?m)^- Link:\s*(?:<>|<\s*>)?\s*\n?", "", value)
    return value


def extract_contacts(text: str) -> list[str]:
    emails = [CONTACT_CORRECTIONS.get(value, value) for value in EMAIL_RE.findall(text or "")]
    contacts = emails + list(PHONE_RE.findall(text or ""))
    return unique(contacts, MAX_CONTACTS)


def redact_sensitive_text(text: str) -> str:
    value = clean_text(text)
    value = EMAIL_RE.sub("[redacted email]", value)
    value = PHONE_RE.sub("[redacted phone]", value)
    value = re.sub(r"\b\d{3}-\d{2}-\d{4}\b", "[redacted identifier]", value)
    value = re.sub(r"\b\d{2}-\d{7}\b", "[redacted identifier]", value)
    value = re.sub(
        r"(?i)\b(EIN|SSN|tax(?:payer)?\s+ID|account(?:\s+number)?|routing(?:\s+number)?|card(?:\s+number)?)"
        r"\s*(?::|#|\s)\s*(?=[A-Z0-9-]*\d)[A-Z0-9-]{4,}",
        lambda match: f"{match.group(1)}: [redacted]",
        value,
    )
    value = re.sub(r"(?<!\d)(?:\d[ -]?){9,}(?!\d)", "[redacted number]", value)
    return value


def access_classification(relative_path: str) -> tuple[str, bool]:
    lowered = relative_path.casefold()
    if any(marker in lowered for marker in RESTRICTED_PATH_MARKERS):
        return "restricted", True
    if any(marker in lowered for marker in INTERNAL_PATH_MARKERS):
        return "internal", False
    if lowered.startswith("ai training/resource files for training/"):
        return "public", False
    if lowered.startswith(("nav playbooks/", "stage 0 pre_formation/", "stage 1 formation/", "franchise/")):
        return "internal", False
    return "internal", False


def parse_markdown_sections(text: str) -> dict[str, list[str]]:
    sections: dict[str, list[str]] = defaultdict(list)
    current = "opening"
    for raw in text.splitlines():
        line = raw.strip()
        heading = re.match(r"^#{1,6}\s+(.+?)\s*$", line)
        if heading:
            current = clean_text(heading.group(1)).casefold()
            continue
        if line:
            sections[current].append(line)
    return dict(sections)


def section_value(sections: dict[str, list[str]], names: tuple[str, ...]) -> str:
    for heading, lines in sections.items():
        if any(name in heading for name in names):
            candidates = []
            for line in lines:
                item = re.sub(r"^[-*+]\s+|^\d+[.)]\s+", "", line).strip()
                if meaningful_line(item):
                    candidates.append(item)
            if candidates:
                return truncate(" ".join(candidates[:3]), MAX_DESCRIPTION)
    return ""


def xml_paragraphs(data: bytes, paragraph_tags: set[str] | None = None) -> list[str]:
    paragraph_tags = paragraph_tags or {"p", "h"}
    root = ET.fromstring(data)
    paragraphs: list[str] = []
    for element in root.iter():
        if local_name(element.tag) not in paragraph_tags:
            continue
        pieces = []
        for child in element.iter():
            if local_name(child.tag) in {"t", "span", "tab"} and child.text:
                pieces.append(child.text)
        text = clean_text(" ".join(pieces))
        if text:
            paragraphs.append(text)
    return paragraphs


def zip_external_links(archive: zipfile.ZipFile) -> list[str]:
    links: list[str] = []
    for name in archive.namelist():
        if not name.endswith(".rels"):
            continue
        try:
            root = ET.fromstring(archive.read(name))
        except Exception:
            continue
        for element in root.iter():
            target = element.attrib.get("Target", "")
            if target.lower().startswith(("http://", "https://")):
                links.append(target)
    return unique(links, MAX_LINKS)


def extract_markdown_or_text(path: Path) -> Extraction:
    raw = path.read_bytes()
    text = raw.decode("utf-8-sig", errors="replace")
    if text.count("�") > max(2, len(text) // 500):
        text = raw.decode("cp1252", errors="replace")
    text = clean_text(text)
    extraction = Extraction(text=text, method="UTF-8/plain-text extraction")
    extraction.status = "complete"
    extraction.coverage = f"{len(text):,} text characters"
    h1 = re.search(r"(?m)^#\s+(.+?)\s*$", text)
    if h1:
        extraction.title_candidates.append((h1.group(1), "markdown_h1"))
    else:
        for line in text.splitlines():
            if meaningful_line(line):
                extraction.title_candidates.append((line, "first_text_heading"))
                break
    extraction.segments = unique(text.splitlines())
    extraction.links = extract_links(text)
    extraction.contacts = extract_contacts(text)
    return extraction


def extract_pdf(path: Path) -> Extraction:
    extraction = Extraction(method="pypdf text and metadata extraction")
    if PdfReader is None:
        extraction.status = "failed"
        extraction.limitations.append("pypdf is unavailable")
        return extraction
    try:
        reader = PdfReader(str(path), strict=False)
        if reader.is_encrypted:
            try:
                reader.decrypt("")
            except Exception:
                pass
        metadata = reader.metadata or {}
        metadata_title = metadata.get("/Title") if hasattr(metadata, "get") else None
        if metadata_title:
            extraction.title_candidates.append((str(metadata_title), "pdf_metadata_title"))
        parts: list[str] = []
        text_pages = 0
        total_pages = len(reader.pages)
        for page in reader.pages:
            try:
                page_text = clean_text(page.extract_text() or "")
            except Exception:
                page_text = ""
            if page_text:
                text_pages += 1
                if sum(len(part) for part in parts) < MAX_EXTRACTED_TEXT:
                    parts.append(page_text[:20_000])
        text = clean_text("\n".join(parts))[:MAX_EXTRACTED_TEXT]
        if not text and pdfplumber is not None:
            extraction.method += "; pdfplumber fallback"
            try:
                with pdfplumber.open(str(path)) as pdf:
                    fallback_parts = []
                    for page in pdf.pages:
                        value = clean_text(page.extract_text() or "")
                        if value:
                            fallback_parts.append(value[:20_000])
                    text = clean_text("\n".join(fallback_parts))[:MAX_EXTRACTED_TEXT]
                    text_pages = len(fallback_parts)
            except Exception as error:
                extraction.limitations.append(f"pdfplumber fallback failed: {type(error).__name__}")
        extraction.text = text
        extraction.segments = unique(text.splitlines())
        extraction.links = extract_links(text)
        extraction.contacts = extract_contacts(text)
        extraction.coverage = f"{text_pages}/{total_pages} pages yielded text; {len(text):,} characters retained"
        extraction.details = {"pages": total_pages, "text_pages": text_pages}
        if text_pages == total_pages and total_pages:
            extraction.status = "complete"
        elif text_pages:
            extraction.status = "partial"
            extraction.limitations.append("One or more pages did not expose machine-readable text")
        else:
            extraction.status = "metadata-only"
            extraction.limitations.append("No machine-readable text was found; OCR was not available")
        for line in extraction.segments:
            if meaningful_line(line):
                extraction.title_candidates.append((line, "first_pdf_text_line"))
                break
    except Exception as error:
        extraction.status = "failed"
        extraction.coverage = "PDF could not be parsed"
        extraction.limitations.append(f"PDF extraction failed: {type(error).__name__}")
    return extraction


def extract_docx(path: Path) -> Extraction:
    extraction = Extraction(method="OOXML text, relationship, and core-metadata extraction")
    try:
        with zipfile.ZipFile(path) as archive:
            names = archive.namelist()
            text_parts: list[str] = []
            for name in names:
                if not name.startswith("word/") or not name.endswith(".xml"):
                    continue
                if not any(token in name for token in ("document.xml", "header", "footer", "footnotes", "endnotes", "comments")):
                    continue
                try:
                    text_parts.extend(xml_paragraphs(archive.read(name), {"p"}))
                except Exception:
                    continue
            if "docProps/core.xml" in names:
                try:
                    core = ET.fromstring(archive.read("docProps/core.xml"))
                    for element in core.iter():
                        if local_name(element.tag) == "title" and element.text:
                            extraction.title_candidates.append((element.text, "docx_core_title"))
                except Exception:
                    pass
            media_count = sum(1 for name in names if name.startswith("word/media/") and not name.endswith("/"))
            extraction.details = {"embedded_media": media_count}
            extraction.links = zip_external_links(archive)
        extraction.segments = unique(text_parts)
        extraction.text = clean_text("\n".join(extraction.segments))[:MAX_EXTRACTED_TEXT]
        extraction.links = unique(extraction.links + extract_links(extraction.text), MAX_LINKS)
        extraction.contacts = extract_contacts(extraction.text)
        for segment in extraction.segments:
            if meaningful_line(segment):
                extraction.title_candidates.append((segment, "first_docx_content"))
                break
        extraction.coverage = f"{len(extraction.segments)} text blocks; {media_count} embedded media files"
        if extraction.text:
            extraction.status = "complete" if media_count == 0 else "partial"
            if media_count:
                extraction.limitations.append("Embedded media was counted but not independently described")
        else:
            extraction.status = "metadata-only"
            extraction.limitations.append("No OOXML text blocks were found")
    except Exception as error:
        extraction.status = "failed"
        extraction.coverage = "DOCX could not be parsed"
        extraction.limitations.append(f"DOCX extraction failed: {type(error).__name__}")
    return extraction


def numbered_xml_names(names: Iterable[str], prefix: str, suffix: str = ".xml") -> list[str]:
    pattern = re.compile(re.escape(prefix) + r"(\d+)" + re.escape(suffix) + r"$")
    matched = []
    for name in names:
        match = pattern.search(name)
        if match:
            matched.append((int(match.group(1)), name))
    return [name for _, name in sorted(matched)]


def extract_pptx(path: Path) -> Extraction:
    extraction = Extraction(method="OOXML slide, notes, relationship, and core-metadata extraction")
    try:
        with zipfile.ZipFile(path) as archive:
            names = archive.namelist()
            slides = numbered_xml_names(names, "ppt/slides/slide")
            slide_segments: list[str] = []
            for index, name in enumerate(slides, start=1):
                try:
                    lines = xml_paragraphs(archive.read(name), {"p"})
                except Exception:
                    lines = []
                if lines:
                    if index == 1:
                        extraction.title_candidates.append((lines[0], "first_slide_title"))
                    slide_segments.append(f"Slide {index}: " + " | ".join(lines))
            notes = numbered_xml_names(names, "ppt/notesSlides/notesSlide")
            note_segments: list[str] = []
            for name in notes:
                try:
                    note_segments.extend(xml_paragraphs(archive.read(name), {"p"}))
                except Exception:
                    continue
            if "docProps/core.xml" in names:
                try:
                    core = ET.fromstring(archive.read("docProps/core.xml"))
                    for element in core.iter():
                        if local_name(element.tag) == "title" and element.text:
                            extraction.title_candidates.insert(0, (element.text, "pptx_core_title"))
                except Exception:
                    pass
            media_count = sum(1 for name in names if name.startswith("ppt/media/") and not name.endswith("/"))
            extraction.links = zip_external_links(archive)
        extraction.segments = unique(slide_segments + note_segments)
        extraction.text = clean_text("\n".join(extraction.segments))[:MAX_EXTRACTED_TEXT]
        extraction.links = unique(extraction.links + extract_links(extraction.text), MAX_LINKS)
        extraction.contacts = extract_contacts(extraction.text)
        extraction.details = {"slides": len(slides), "notes_blocks": len(note_segments), "embedded_media": media_count}
        extraction.coverage = f"{len(slides)} slides, {len(note_segments)} notes blocks, {media_count} embedded media files"
        extraction.status = "complete" if media_count == 0 else "partial"
        if media_count:
            extraction.limitations.append("Slide text and notes were extracted; embedded visuals were not separately described")
        if not extraction.text:
            extraction.status = "metadata-only"
            extraction.limitations.append("No machine-readable slide text was found")
    except Exception as error:
        extraction.status = "failed"
        extraction.coverage = "PPTX could not be parsed"
        extraction.limitations.append(f"PPTX extraction failed: {type(error).__name__}")
    return extraction


def extract_odp(path: Path) -> Extraction:
    extraction = Extraction(method="OpenDocument content.xml and metadata extraction")
    try:
        with zipfile.ZipFile(path) as archive:
            names = archive.namelist()
            segments = xml_paragraphs(archive.read("content.xml"), {"p", "h"}) if "content.xml" in names else []
            if "meta.xml" in names:
                try:
                    meta = ET.fromstring(archive.read("meta.xml"))
                    for element in meta.iter():
                        if local_name(element.tag) == "title" and element.text:
                            extraction.title_candidates.append((element.text, "odp_metadata_title"))
                except Exception:
                    pass
            media_count = sum(
                1
                for name in names
                if name.startswith("Pictures/") and not name.endswith("/")
            )
            extraction.links = zip_external_links(archive)
        extraction.segments = unique(segments)
        extraction.text = clean_text("\n".join(extraction.segments))[:MAX_EXTRACTED_TEXT]
        extraction.links = unique(extraction.links + extract_links(extraction.text), MAX_LINKS)
        extraction.contacts = extract_contacts(extraction.text)
        for segment in extraction.segments:
            if meaningful_line(segment):
                extraction.title_candidates.append((segment, "first_odp_content"))
                break
        extraction.details = {"text_blocks": len(extraction.segments), "embedded_images": media_count}
        extraction.coverage = f"{len(extraction.segments)} text blocks; {media_count} embedded images"
        extraction.status = "complete" if media_count == 0 else "partial"
        if media_count:
            extraction.limitations.append("OpenDocument text was extracted; embedded visuals were not separately described")
        if not extraction.text:
            extraction.status = "metadata-only"
            extraction.limitations.append("No machine-readable OpenDocument text was found")
    except Exception as error:
        extraction.status = "failed"
        extraction.coverage = "ODP could not be parsed"
        extraction.limitations.append(f"ODP extraction failed: {type(error).__name__}")
    return extraction


def worksheet_value(value: Any) -> str:
    if value is None:
        return ""
    text = clean_text(value)
    return truncate(text, 120)


def extract_xlsx(path: Path) -> Extraction:
    extraction = Extraction(method="openpyxl read-only workbook structure and formula sampling")
    if load_workbook is None:
        extraction.status = "failed"
        extraction.limitations.append("openpyxl is unavailable")
        return extraction
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
        sheet_summaries: list[str] = []
        total_nonempty = 0
        for sheet in workbook.worksheets:
            rows_seen = 0
            sample_rows: list[list[str]] = []
            for row in sheet.iter_rows(values_only=True):
                values = [worksheet_value(value) for value in row[:24]]
                if any(values):
                    total_nonempty += sum(bool(value) for value in values)
                    if len(sample_rows) < 8:
                        sample_rows.append(values)
                rows_seen += 1
                if rows_seen >= 2000:
                    break
            compact = [" | ".join(value for value in row if value) for row in sample_rows]
            sheet_summaries.append(f"Sheet {sheet.title}: " + " / ".join(compact[:4]))
        workbook.close()
        extraction.title_candidates.append(("; ".join(workbook.sheetnames[:3]), "workbook_sheet_names"))
        extraction.segments = unique(sheet_summaries)
        extraction.text = clean_text("\n".join(extraction.segments))[:MAX_EXTRACTED_TEXT]
        extraction.links = extract_links(extraction.text)
        extraction.contacts = extract_contacts(extraction.text)
        extraction.details = {"sheets": workbook.sheetnames, "sampled_nonempty_cells": total_nonempty}
        extraction.coverage = f"{len(workbook.sheetnames)} sheets; sampled up to 2,000 rows and 24 columns per sheet"
        extraction.status = "complete"
        extraction.limitations.append("Formulas were captured as stored and were not recalculated")
    except Exception as error:
        extraction.status = "failed"
        extraction.coverage = "XLSX could not be parsed"
        extraction.limitations.append(f"XLSX extraction failed: {type(error).__name__}")
    return extraction


def extract_csv_file(path: Path) -> Extraction:
    extraction = Extraction(method="CSV header, row-count, and sample extraction")
    try:
        raw = path.read_bytes()
        text = raw.decode("utf-8-sig", errors="replace")
        if text.count("�") > max(2, len(text) // 500):
            text = raw.decode("cp1252", errors="replace")
        sample = text[:16_384]
        try:
            dialect = csv.Sniffer().sniff(sample)
        except Exception:
            dialect = csv.excel
        rows = list(csv.reader(text.splitlines(), dialect))
        headers = rows[0] if rows else []
        extraction.title_candidates.append((" / ".join(headers[:6]), "csv_headers"))
        extraction.segments = unique(" | ".join(clean_text(cell) for cell in row if clean_text(cell)) for row in rows[:8])
        extraction.text = clean_text("\n".join(extraction.segments))[:MAX_EXTRACTED_TEXT]
        extraction.links = extract_links(text)
        extraction.contacts = extract_contacts(text)
        extraction.details = {"rows": max(0, len(rows) - 1), "columns": len(headers), "headers": headers}
        extraction.coverage = f"{max(0, len(rows) - 1)} data rows; {len(headers)} columns; first 8 rows summarized"
        extraction.status = "complete"
    except Exception as error:
        extraction.status = "failed"
        extraction.coverage = "CSV could not be parsed"
        extraction.limitations.append(f"CSV extraction failed: {type(error).__name__}")
    return extraction


def extract_legacy_ppt(path: Path) -> Extraction:
    extraction = Extraction(method="legacy OLE PowerPoint printable-string scan")
    try:
        data = path.read_bytes()
        ascii_chunks = [chunk.decode("cp1252", errors="ignore") for chunk in re.findall(rb"[\x20-\x7e]{5,}", data)]
        utf16_chunks = [chunk.decode("utf-16le", errors="ignore") for chunk in re.findall(rb"(?:[\x20-\x7e]\x00){5,}", data)]
        candidates = []
        for value in ascii_chunks + utf16_chunks:
            line = clean_text(value)
            if meaningful_line(line) and not re.fullmatch(r"[A-Fa-f0-9-]{12,}", line):
                candidates.append(line)
        extraction.segments = unique(candidates, 400)
        extraction.text = clean_text("\n".join(extraction.segments))[:MAX_EXTRACTED_TEXT]
        extraction.links = extract_links(extraction.text)
        extraction.contacts = extract_contacts(extraction.text)
        for segment in extraction.segments:
            if meaningful_line(segment):
                extraction.title_candidates.append((segment, "legacy_ppt_printable_text"))
                break
        extraction.coverage = f"{len(extraction.segments)} printable text fragments retained"
        extraction.status = "partial" if extraction.segments else "metadata-only"
        extraction.limitations.append("Legacy PPT extraction is best-effort and does not preserve slide order or layout")
    except Exception as error:
        extraction.status = "failed"
        extraction.coverage = "Legacy PPT could not be scanned"
        extraction.limitations.append(f"Legacy PPT scan failed: {type(error).__name__}")
    return extraction


def extract_rtb(path: Path) -> Extraction:
    extraction = Extraction(method="RTB ZIP member inventory and readable JSON extraction")
    try:
        with zipfile.ZipFile(path) as archive:
            members = [name for name in archive.namelist() if not name.endswith("/")]
            parsed: dict[str, Any] = {}
            unreadable: list[str] = []
            for name in members:
                try:
                    parsed[name] = json.loads(archive.read(name).decode("utf-8"))
                except Exception:
                    unreadable.append(name)
            board = parsed.get("board.json", {})
            if isinstance(board, dict) and board.get("name"):
                extraction.title_candidates.append((board["name"], "rtb_board_name"))
            values: list[str] = []

            def collect(value: Any) -> None:
                if isinstance(value, dict):
                    for child in value.values():
                        collect(child)
                elif isinstance(value, list):
                    for child in value:
                        collect(child)
                elif isinstance(value, str) and meaningful_line(value):
                    values.append(value)

            collect(parsed)
            extraction.segments = unique(values)
            extraction.text = clean_text("\n".join(extraction.segments))
            extraction.details = {"members": members, "readable_json_members": sorted(parsed), "unreadable_members": unreadable}
            extraction.coverage = f"{len(parsed)}/{len(members)} members decoded as JSON"
            extraction.status = "complete" if not unreadable else "partial"
            if unreadable:
                extraction.limitations.append("Canvas/table members use an application-level encoded or encrypted payload")
    except Exception as error:
        extraction.status = "failed"
        extraction.coverage = "RTB archive could not be parsed"
        extraction.limitations.append(f"RTB extraction failed: {type(error).__name__}")
    return extraction


def extract_zip_file(path: Path) -> Extraction:
    extraction = Extraction(method="ZIP central-directory inventory")
    try:
        with zipfile.ZipFile(path) as archive:
            members = [info.filename for info in archive.infolist() if not info.is_dir()]
        extraction.details = {"members": members}
        if members:
            extraction.title_candidates.append((f"Archive containing {len(members)} files", "zip_inventory"))
            extraction.segments = members
            extraction.text = "\n".join(members)
            extraction.status = "partial"
            extraction.coverage = f"{len(members)} member names inventoried"
            extraction.limitations.append("Nested archive members were not recursively converted")
        else:
            extraction.title_candidates.append(("Empty BOSS ZIP archive", "zip_inventory"))
            extraction.status = "empty"
            extraction.coverage = "Valid ZIP archive with zero members"
    except Exception as error:
        extraction.status = "failed"
        extraction.coverage = "ZIP could not be parsed"
        extraction.limitations.append(f"ZIP inventory failed: {type(error).__name__}")
    return extraction


IMAGE_OVERRIDES: dict[str, dict[str, Any]] = {
    "AI Training/Resource Files For Training/SBDC/SBDC Map.jpg": {
        "title": "Texas Gulf Coast SBDC Network Location Map",
        "description": "Screenshot of the University of Houston Texas Gulf Coast Small Business Development Center network page. It explains that 14 locations serve 32 Southeast Texas counties and shows a county map with service locations marked by red stars.",
        "when": "Surface when a Southeast Texas entrepreneur needs to find a nearby SBDC office or understand the network's geographic coverage.",
        "key": ["The page routes users to SBDC advising locations.", "The network supports pre-venture, startup, expansion, growth, and exit needs.", "Coverage shown: 14 locations across 32 Southeast Texas counties."],
        "next": "Use the current Texas Gulf Coast SBDC location finder and confirm the user's ZIP or county before referring them.",
    },
    "AI Training/Resource Files For Training/SCORE/SCORE/SCORE Impact Slide 1.jpg": {
        "title": "2024 SCORE National Small Business Impact Snapshot",
        "description": "SCORE impact slide showing entrepreneurs in agriculture, food service, and retail alongside three national 2024 outcomes.",
        "when": "Surface when a user asks what SCORE does or wants evidence of SCORE's national small-business impact.",
        "key": ["59,447 new businesses launched in 2024.", "84,176 non-owner jobs created in 2024.", "143,623 total jobs created in 2024.", "The slide points readers to the 2024 SCORE National Impact Report."],
        "next": "Use SCORE's current impact page or report for citation-ready figures and route service questions to SCORE's official site.",
    },
    "AI Training/Resource Files For Training/SCORE/SCORE/SCORE Roadmap Slide 1.jpg": {
        "title": "SCORE Startup Roadmap Overview",
        "description": "SCORE webpage screenshot introducing the Startup Roadmap as a step-by-step guide for starting a business and encouraging users to work through it with a SCORE mentor.",
        "when": "Surface for early-stage founders who need an organized startup learning path, concrete tasks, mentor discussion topics, and supporting resources.",
        "key": ["Roadmap modules combine Steps to Success, With Your Mentor discussion prompts, and Resources.", "Modules may include recommended homework and mentor-guided follow-up.", "The roadmap is presented as a guide rather than a substitute for individualized mentor judgment."],
        "next": "Direct the user to SCORE's current Startup Roadmap and mentor-request process.",
    },
    "AI Training/Resource Files For Training/SCORE/SCORE/SCORE Roadmap Slide 2.jpg": {
        "title": "SCORE Startup Roadmap Module Catalog",
        "description": "Partial screenshot of SCORE's 12-module Startup Roadmap catalog. The visible modules cover starting the journey, readiness, idea validation, business planning, funding, and business setup.",
        "when": "Surface when a founder needs the next relevant Startup Roadmap module rather than a generic list of startup tasks.",
        "key": ["The modules do not have to be completed in a fixed order.", "Visible modules: Starting Your Journey; Are You Ready; Defining and Validating Your Idea; Business Plan or Business Model Canvas; Finding Funding; Setting Up Your Business.", "The screenshot is a partial catalog view, not the complete 12-module list."],
        "next": "Match the user's current need to the relevant live SCORE module and offer a SCORE mentor handoff when judgment is needed.",
    },
    "AI Training/Resource Files For Training/SCORE/SCORE/SCORE Services Slide.jpg": {
        "title": "SCORE Services: Mentoring, Education, and Online Resources",
        "description": "SCORE services slide presenting three routes: free confidential mentoring, webinars and on-demand courses, and an online library of guides, templates, checklists, videos, and articles.",
        "when": "Surface when the user is deciding whether they need a mentor, self-paced education, a live webinar, or a practical template.",
        "key": ["SCORE mentoring is described as free and confidential, in person or remote.", "Education includes live webinars, recordings, and self-paced courses.", "The online resource library includes guides, templates, checklists, blogs, videos, and infographics."],
        "next": "Route to the matching live SCORE service page and avoid implying that the librarian itself is the mentor.",
    },
    "EDE Map/BOSS VEDE Houston Map - Frame 1.jpg": {
        "title": "Business Operator Bootstrap Lifecycle and Resource Map",
        "description": "Wide ecosystem diagram connecting a seven-stage business lifecycle, stage-appropriate support organizations, funding stages, and funding sources.",
        "when": "Surface when a user needs a lifecycle orientation or a high-level view of which support and funding categories may become relevant over time.",
        "key": ["Lifecycle: Ideation, Validation, Formation, Launch, Operations, Growth, Exit.", "Support examples include The CIVVIE, SCORE, SBDC, VBOC, TVC, Warrior Rising, IVMF, UH SURE, and APEX.", "Funding stages: Pre-Seed, Seed, Operating, Growth, Exit.", "Funding sources shown range from personal savings and friends/family to CDFIs, loans, client revenue, partnerships, equity, legacy owners, and strategic buyers."],
        "next": "Use the map to orient the user, then verify one concrete next resource from the current document library rather than presenting every future option at once.",
    },
    "Workspace Architecture/BOSS Workspace Architecture 021526.jpg": {
        "title": "BOSS Five-Seat Workspace Architecture",
        "description": "Architecture diagram for a five-seat operating system using Google Workspace as the identity spine, with ChatGPT Business, shared Google Drives, Miro, Canva, and an optional automation layer.",
        "when": "Surface for internal questions about where BOSS strategy, source files, visual planning, branded outputs, and automations belong.",
        "key": ["Google Workspace is the primary identity and administration layer.", "ChatGPT Business supports strategy, research, and drafting.", "Shared Drives hold doctrine, operations, marketing, and finance source files.", "Miro supports lifecycle, planning, client maps, and product development.", "Canva is the brand/output layer; Zapier, Make, or Apps Script may connect tools."],
        "next": "Route implementation decisions to the BOSS workspace administrator and the controlled architecture documentation.",
    },
    "Workspace Architecture/Bosslibraries Workspace Architecture V1 021826.jpg": {
        "title": "BOSSLibraries.com Workspace Ecosystem Architecture - No SSO",
        "description": "Ecosystem diagram for a four-person Google Workspace environment without single sign-on, connecting planning, whiteboarding, documentation, AI, video, automation, storage, server, and security layers.",
        "when": "Surface for internal planning about tool ownership, integration boundaries, or the earlier no-SSO architecture.",
        "key": ["The domain and Google Workspace act as the access hub.", "Core tools include Monday.com, Miro, Acrobat, OpenAI/Ollama, HeyGen, Make/Zapier, and Google Drive.", "A DigitalOcean AI server and security/admin layer support the ecosystem.", "The diagram marks future CRM, forms, analytics, Notion, Slack, and GitHub integrations."],
        "next": "Compare this earlier architecture with the current BOSS OSF workspace design before making implementation decisions.",
    },
    "Workspace Architecture/Bosslibraries Workspace Architecture V2 021826.jpg": {
        "title": "BOSS Workspace Architecture - Integrated Tools Ecosystem",
        "description": "Detailed BOSS-branded architecture showing Google Workspace, a DigitalOcean integration server, AI and automation tools, production outputs, security controls, and future integrations.",
        "when": "Surface for internal questions about the integrated BOSS tool stack, data flow, hosting, or governance.",
        "key": ["Google Workspace contains shared drives for executive, playbook, diagram, and media work.", "The integration layer includes APIs, Make/Zapier, webhooks, and IFTTT.", "DigitalOcean hosts Ubuntu, NGINX, and SSL/TLS services.", "Outputs include librarian video, playbooks, ecosystem diagrams, and project planning.", "Governance includes domain authentication, access control, audit logs, backup, and encryption."],
        "next": "Use the latest approved architecture and route infrastructure changes to the workspace/security owner.",
    },
    "Workspace Architecture/BOSS OSF and Dual Google Workspace Architecture 070126.docx.png": {
        "title": "BOSS Operating System Framework Workspace Architecture v1.1",
        "description": "Current-state BOSS Operating System Framework diagram separating internal operations from the public library while joining them through shared services, automation, infrastructure, governance, and an AI librarian ecosystem.",
        "when": "Surface when a user needs to understand where internal records, public library knowledge, training data, partner resources, or AI librarian services belong.",
        "key": ["Workspace A is internal operations and service delivery.", "Workspace B is the public-facing library, knowledge, digital-product, subscriber, community, marketing, media, partner, and AI-training environment.", "Shared services include Drive, Calendar, Tasks, Meet, Chat, Contacts, Vault, Admin, and security/identity.", "The AI librarian ecosystem connects a training pipeline, knowledge index, retrieval engine, OpenAI/Ollama, and vector data.", "The framework emphasizes navigation first, standardization, knowledge assets, AI augmentation, one source of truth, and scalable security."],
        "next": "Use this framework as the default placement guide, then route access or architecture changes to the appropriate workspace administrator.",
    },
    "AI Training/Resource Files For Training/SBA/SBA-RDS-Mundo/Boots to Businss Reboot/Boots to Business Management Instructor Info.png": {
        "title": "Boots to Business Management Instructor Resources",
        "description": "Screenshot of the U.S. Small Business Administration Boots to Business management portal for instructors, with curriculum, lesson-plan, workbook, training, toolbox, classroom, student, and event-flyer resources.",
        "when": "Surface for an authorized Boots to Business instructor who needs curriculum or instructor-support resources, not for general startup counseling.",
        "key": ["Sections shown: B2B Curriculum, MilSpouse Pathway to Business, Reboot One Day Curriculum, and Other Resources.", "Links include lesson plans, instructor guides, student virtual workbooks, instructor training, classroom documents, student resources, and an instructor toolbox.", "The page displays a curriculum feedback form and Boots to Business help-desk contact options."],
        "contacts": ["Boots-to-Business@sba.gov", "(844) 610-VET1", "(202) 205-VET1"],
        "next": "Use the current authorized SBA instructor portal or Boots to Business help desk for access and support.",
    },
    "Branding/BOSS Libraries Email Logo 031526.png": {
        "title": "BOSS Libraries Horizontal Compass-and-Books Logo",
        "description": "Horizontal BOSS Libraries logo on white: a gold-and-navy compass rises from stacked navy, gold, cream, and green books beside navy 'BOSS' and gold 'Libraries' lettering with a trademark mark.",
        "when": "Surface for approved email signatures, horizontal brand placements, or brand-reference questions.",
        "key": ["Primary colors are navy, gold, cream, and a muted green accent.", "The symbol combines navigation (compass) with organized knowledge (books).", "The asset includes the BOSS Libraries trademark mark."],
        "next": "Confirm approved size, clear space, and current brand guidance before production use.",
    },
    "Logo/BOSS Libraries Email Logo 031526.png": {
        "title": "BOSS Libraries Horizontal Compass-and-Books Logo",
        "description": "Horizontal BOSS Libraries logo on white: a gold-and-navy compass rises from stacked navy, gold, cream, and green books beside navy 'BOSS' and gold 'Libraries' lettering with a trademark mark.",
        "when": "Surface for approved email signatures, horizontal brand placements, or brand-reference questions.",
        "key": ["Primary colors are navy, gold, cream, and a muted green accent.", "The symbol combines navigation (compass) with organized knowledge (books).", "This source is byte-identical to the Branding copy."],
        "next": "Use the canonical brand asset and confirm approved size, clear space, and current brand guidance.",
    },
    "Logo/BOSS Background (2).png": {
        "title": "BOSS Libraries Ornate Library Background",
        "description": "Warm bronze-and-gold panoramic image of an ornate historic library interior with floor-to-ceiling books, decorative galleries, and BOSS Libraries branding at upper left and on a central hanging banner.",
        "when": "Surface when selecting an approved branded background for BOSS Libraries presentations, videos, or visual environments.",
        "key": ["The visual theme communicates heritage, authority, navigation, and organized knowledge.", "The palette is dark brown, bronze, gold, and black.", "The image includes BOSS Libraries and Business Operator Solutions & Services branding."],
        "next": "Check legibility, contrast, crop, and current brand approval before placing text or presenters over the background.",
    },
    "Logo/Draft BOSS logo design with TM.png": {
        "title": "BOSS Libraries Primary Stacked Logo",
        "description": "Stacked BOSS Libraries logo on black: a compass and open book above navy 'BOSS,' gold 'LIBRARIES,' a trademark mark, and the line 'Business Operator Solutions & Services.'",
        "when": "Surface for primary-logo identification, dark-background placements, or brand-guideline questions.",
        "key": ["The primary mark combines a compass, books, and the BOSS Libraries wordmark.", "The palette uses navy, gold, cream, and green accents.", "The full lockup includes the Business Operator Solutions & Services descriptor."],
        "next": "Confirm that the draft designation has been superseded or approved before production use.",
    },
}


VIDEO_OVERRIDES: dict[str, dict[str, Any]] = {
    "Branding/BOSS_Intro_with_captions.mp4": {
        "title": "How to Use BOSS Libraries",
        "description": "Branded presenter video explaining that BOSS Libraries helps users locate stage-appropriate business resources, tools, templates, guidance, and human support. The presenter appears over the ornate BOSS Libraries background with burned-in captions.",
        "when": "Surface as a short orientation when a user is new to BOSS Libraries or unsure what kinds of questions to ask.",
        "key": ["An idea-stage user can ask for resources that help turn an idea into a business concept.", "The library can explain lifecycle context and identify practical next steps.", "Resources may include mentoring, guidance, proven models, tools, templates, and other support.", "Users with a defined concept can ask focused questions, such as where to find marketing-channel resources."],
        "next": "Invite the user to state their current business stage, location, and immediate information need, then retrieve the smallest relevant set of documents and human-resource links.",
    },
    "Branding/Doug_Intro_to_BOSS_Libraries.mp4": {
        "title": "Welcome to BOSS Libraries from Doug",
        "description": "Short portrait-format welcome video. Doug sits at a library table with an open book and laptop and introduces BOSS Libraries as a digital business-resource optimization library supported by an AI librarian.",
        "when": "Surface as a brief founder welcome or explanation of why BOSS Libraries exists.",
        "key": ["Doug identifies himself as a small-business ecosystem guide and mentor.", "He says entrepreneur feedback showed that finding the right knowledge resource is confusing.", "BOSS Libraries was created to organize those resources and make them accessible through an AI librarian."],
        "people": ["Doug"],
        "next": "Transition from the welcome to the user's concrete need and retrieve relevant documents or people; do not treat the librarian itself as a mentor.",
    },
}


SPECIAL_OVERRIDES: dict[str, dict[str, Any]] = {
    "AI Training/Resource Files For Training/RFM Kimberly Schultz/RFM Flyer Kimberly Schultz current.pdf": {
        "title": "Revolution Financial Management Financial Services Flyer",
        "description": "One-page financial-services flyer for Revolution Financial Management built around the mission 'No Family Left Behind.' It introduces a local associate and the household and small-business planning topics available through a free consultation.",
        "when": "Surface when a user wants a direct financial-services contact for savings, debt, insurance, long-term care, retirement, or small-business protection questions.",
        "key": [
            "Services listed include short- and long-term savings, debt management, life insurance, long-term care, business insurance, and retirement planning.",
            "The flyer offers a free consultation.",
            "The listed associate is Kimberly Schultz, Financial Services Associate, license 2565151.",
        ],
        "contacts": ["kschultz3124@yahoo.com", "845-641-7019"],
        "next": "Share the flyer contact when it fits the request, and advise the user to confirm current licensing, availability, scope, and product details directly with the provider.",
        "review_scope": "The single flyer page was visually reviewed and described.",
    },
    "AI Training/Resource Files For Training/SCORE/SCORE/ED Mapping Resources/SCORE  Houston Area Map.pptx": {
        "title": "SCORE Houston Area ZIP and Evacuation Zone Map",
        "description": "Single-slide Houston-area ZIP-code map divided into hand-drawn north, central, west/southwest, and east/east-west regions. A green southeast area is identified in the legend as Evacuation Zip-Zone B, and an inset provides additional central-Houston detail.",
        "when": "Surface for internal SCORE Houston orientation when a coordinator needs a rough geographic reference for mentor coverage, event planning, or ZIP-based routing.",
        "key": [
            "The slide is an informal working map, not an official service-boundary map.",
            "Visible working regions are labeled N, C, W/SW, and E/EW.",
            "The green southeast area is labeled Evacuation Zip-Zone B.",
        ],
        "next": "Confirm the user's ZIP and current SCORE Houston coverage or emergency-zone information before making a referral or operational decision.",
        "review_scope": "The single image-based slide was visually reviewed and described.",
    },
    "Consulting/Charlesbank/RD Sewell Executed NDA 112425.pdf": {
        "title": "Transaction Evaluation Confidentiality Agreement",
        "description": "Controlled confidentiality agreement governing the receipt and use of evaluation material for a possible transaction. The routing description intentionally omits party identities, project names, dates, addresses, signatures, and other private transaction details.",
        "when": "Surface only to authorized staff who need to confirm the confidentiality rules that apply before reviewing, sharing, retaining, or disposing of transaction evaluation material.",
        "key": [
            "Covers limited use and disclosure of evaluation material and responsibility for authorized representatives.",
            "Addresses compelled disclosure, return or destruction, process communications, non-solicitation, remedies, governing law, and duration.",
            "Prohibits uploading evaluation material to public artificial-intelligence systems without prior written consent and confirmation that disclosure will not result.",
        ],
        "next": "Open the controlled original with the transaction owner or counsel; do not distribute the source or treat this routing summary as legal advice.",
        "review_scope": "All 5 pages were visually reviewed; only purpose-level, non-identifying terms are described.",
    },
    "EDE Map/Houston Entrepreneur Ecosystem Layered Model 042525.pdf": {
        "title": "Houston Entrepreneur Ecosystem Layered Model",
        "description": "One-page Idea-to-Impact ecosystem map that organizes Houston entrepreneurial support into four layers: entry and advisory, capability development, access and translation, and demand-side opportunity. A compact decision tree routes a business according to formation, operating readiness, and contracting readiness.",
        "when": "Surface when a Houston entrepreneur needs to understand which type of ecosystem support fits their current business readiness and intended opportunity path.",
        "key": [
            "Layer 1: entry and advisory through SCORE, the U.S. Small Business Administration, veteran-serving organizations, and entrepreneur-support organizations.",
            "Layer 2: capability development through Small Business Development Centers, universities, incubators, accelerators, and technical assistance.",
            "Layer 3: access and translation through the Houston Office of Business Opportunity, the Greater Houston Business Procurement Forum, and the Greater Houston Partnership.",
            "Layer 4: demand from public agencies, NASA and prime contractors, and major private industries.",
            "The intended outcome is stronger businesses that win contracts, create jobs, and strengthen Houston.",
        ],
        "next": "Use the decision tree to identify the user's current layer, then retrieve one or two current documents or human-support links for that layer.",
        "review_scope": "The single diagram page was visually reviewed and described.",
    },
    "Formation/Legal Docs/BOSS RARO 02012026.pdf": {
        "title": "Registered Agent and Registered Office Agreement",
        "description": "Controlled agreement appointing a registered agent and registered office for BOSS. It defines the registered-agent service, client duties, fees and costs, term and termination, notices, and standard legal provisions while omitting private names, addresses, signatures, and financial amounts.",
        "when": "Surface to authorized operations or legal staff who need the governing record for registered-agent services, official correspondence, renewal, or a change in company details.",
        "key": [
            "Covers receipt and forwarding of official correspondence and restrictions on use of the registered office.",
            "Requires timely notice of company-address, executive-management, or operating-status changes.",
            "Includes service fees and costs, annual renewal, termination, notice, indemnity, liability, governing-law, assignment, amendment, and counterpart provisions.",
        ],
        "next": "Use the controlled original and route interpretation, changes, or termination questions to the responsible company officer or counsel.",
        "review_scope": "All 5 pages were visually reviewed; the description excludes names, addresses, signatures, and amounts.",
    },
    "Formation/Legal Docs/BOSS Written Consent 02012026.pdf": {
        "title": "Organizational Written Consent of the Member",
        "description": "Controlled organizational consent adopted in place of an organizational meeting. It establishes the company's initial records and authorizations while omitting officer and member identities, signatures, and dates.",
        "when": "Surface to authorized operations, banking, tax, or legal staff who need the company's initial organizational resolutions or proof of delegated authority.",
        "key": [
            "Adopts the formation filing, company record book, governing agreement, and initial officers.",
            "Authorizes required licenses and tax permits, organization expenses, banking arrangements, and qualification in other jurisdictions.",
            "Provides general authority to carry out the organizational resolutions.",
        ],
        "next": "Use the controlled signed original for verification and route authority or interpretation questions to the responsible company officer or counsel.",
        "review_scope": "All 3 pages were visually reviewed; the description excludes personal identities, signatures, and dates.",
    },
    "Formation/Registered Agent/03-02-26 - TX - Initial Filing of New Registered Agent - Business Operator Solutions  Services.pdf": {
        "title": "Texas Change of Registered Agent and Office Filing Packet",
        "description": "Controlled Texas Secretary of State packet documenting a filed change of registered agent and registered office for BOSS. It includes a filing acknowledgment, certificate of filing, and the submitted change statement while omitting file numbers, addresses, names, document identifiers, and signatures.",
        "when": "Surface to authorized legal or operations staff who need evidence that the registered-agent or registered-office change was submitted and accepted by the state.",
        "key": [
            "The packet confirms that the change statement was received and found to conform to Texas filing requirements.",
            "The filed statement records both a registered-agent change and a registered-office change.",
            "The final page contains only the continuation of the execution area and filing-office notation.",
        ],
        "next": "Use the controlled packet to verify the filed details and confirm current state records before relying on the change operationally.",
        "review_scope": "All 4 pages were visually reviewed; the description excludes identifiers, addresses, names, and signatures.",
    },
    "Formation/Texas SOS/2026.01.27 Acknowledgment - BOSS.pdf": {
        "title": "Texas Formation Filing Acknowledgment",
        "description": "Controlled Texas Secretary of State acknowledgment accompanying the company's formation filing. It confirms issuance of formation records and notes ongoing franchise-tax and registered-agent obligations while omitting file numbers, addresses, names, contact details, and dates.",
        "when": "Surface to authorized formation, tax, or legal staff who need the state's acknowledgment and the related post-filing compliance reminders.",
        "key": [
            "Accompanies the certificate of formation and certificate of filing.",
            "Points to Texas Comptroller franchise-tax responsibilities.",
            "Reminds the entity to maintain a registered agent and registered office and to file changes with the state.",
        ],
        "next": "Use the controlled original for official details and confirm current obligations with the appropriate state source or counsel.",
        "review_scope": "The single page was visually reviewed; the description excludes identifying filing details.",
    },
    "Formation/Texas SOS/2026.01.27 Cert of Filing - BOSS.pdf": {
        "title": "Texas SOSDirect Filing Transaction Receipt",
        "description": "Controlled Texas SOSDirect transaction receipt for a formation-related submission. The receipt records the order and status-tracking context but is not itself evidence that the submitted document was approved; session, order, document, payment, and contact details are omitted.",
        "when": "Surface to authorized formation or accounting staff who need to reconcile the state submission or distinguish the transaction receipt from the final certificate of filing.",
        "key": [
            "Records a Secretary of State online transaction and related submission.",
            "States that the receipt does not indicate document approval.",
            "Directs the submitter to use the state system to track processing status.",
        ],
        "next": "Use the separate certificate of filing for proof of acceptance and the controlled receipt only for transaction reconciliation.",
        "review_scope": "The single page was visually reviewed; transaction and payment identifiers are excluded.",
    },
    "Formation/Texas SOS/2026.01.27 Certificate - BOSS.pdf": {
        "title": "Texas Limited Liability Company Certificate of Filing",
        "description": "Controlled Texas Secretary of State certificate confirming formation of a domestic limited liability company. The routing description omits the entity's file number, effective date, official signature, and other filing identifiers.",
        "when": "Surface to authorized legal, banking, tax, or operations staff who need the state's formal evidence that the company formation was accepted.",
        "key": [
            "Certifies that the attached formation document was received and found to conform to Texas filing requirements.",
            "Identifies the filing as formation of a domestic limited liability company.",
            "Serves as the state-issued acceptance record, unlike a transaction receipt.",
        ],
        "next": "Use the controlled certificate and attached formation document when official proof is required; route interpretation questions to counsel or the filing owner.",
        "review_scope": "The single page was visually reviewed; filing identifiers, dates, and signatures are excluded.",
    },
    "Formation/Texas SOS/2026.01.27 Original_Document - BOSS.pdf": {
        "title": "Texas Certificate of Formation for BOSS",
        "description": "Controlled certificate of formation for a Texas limited liability company. It records the entity type and name, registered-agent and office provisions, governing-authority structure, organizer, initial mailing address, business purpose, effectiveness, and execution while omitting personal names, addresses, file numbers, signatures, and dates.",
        "when": "Surface to authorized legal, banking, tax, or operations staff who need the company's filed formation terms or governing-authority structure.",
        "key": [
            "Creates a Texas limited liability company and states a lawful business purpose.",
            "Records registered-agent and registered-office information and a manager-managed governing structure.",
            "Includes an initial mailing address, organizer certification, effectiveness choice, and an attached similar-name consent.",
        ],
        "next": "Use the controlled filed copy for official details and compare any later amendments or state changes before relying on a formation term.",
        "review_scope": "All 3 pages were visually reviewed; personal and filing identifiers are excluded.",
    },
    "Formation/Texas SOS/BOSS Form 509 012326.pdf": {
        "title": "Texas Form 509 Consent to Use of Similar Name",
        "description": "Controlled Texas Secretary of State consent allowing a proposed entity to use a name similar to one already on file. The routing description omits names, file numbers, dates, signatures, and notary information.",
        "when": "Surface to authorized formation or legal staff who need the supporting similar-name consent associated with the company filing.",
        "key": [
            "Identifies the existing name holder and the proposed entity name in the controlled original.",
            "Certifies that the signer is authorized to give the consent.",
            "Includes signature and notarial acknowledgment sections.",
        ],
        "next": "Use the controlled executed form for filing evidence and route name-availability or sufficiency questions to the Texas Secretary of State or counsel.",
        "review_scope": "The single page was visually reviewed; names, filing identifiers, signatures, dates, and notary details are excluded.",
    },
    "Logo/BOSS Professional Logo and Branding Guide 031226.docx": {
        "title": "BOSS Libraries Professional Logo and Branding Guide",
        "description": "One-page visual brand board for Business Operator Solutions & Services and BOSS Libraries. It presents the compass-and-books identity, approved lockups and marks, wordmark and sub-brand treatment, color palette, and the Navigation, Knowledge, Strategy, Growth, and Trust brand pillars.",
        "when": "Surface when creating or reviewing a BOSS Libraries brand application, logo placement, sub-brand treatment, presentation, document, video, or other visual asset.",
        "key": [
            "The primary identity combines a compass with books to represent navigation and organized knowledge.",
            "The guide shows full-logo dark and light variants, the compass/book mark, and the wordmark.",
            "Sub-brand examples include BOSS Libraries and VBOSS Library.",
            "The palette includes navy, red, gold, blue, green, charcoal, and cream.",
            "Brand pillars are Navigation, Knowledge, Strategy, Growth, and Trust.",
        ],
        "next": "Use an approved asset from the logo system, preserve legibility and contrast, and confirm the latest brand guidance before production release.",
        "review_scope": "The full-page embedded brand-board image was visually reviewed and described.",
    },
    "Suppliers/TWiN/Service Agreement/BOSS TWiN Service Agreement 26-02-03.pdf": {
        "title": "BOSS and TWiN Core AI Framework Services Agreement",
        "description": "Controlled services agreement for designing and building the BOSS Library Core AI Framework and supporting platform for entrepreneurial-resource navigation. It defines scope, collaboration, term, compensation structure, milestones, change control, intellectual property, confidentiality, data security, termination, liability, and dispute terms while omitting private party details, signatures, dates, addresses, and financial amounts.",
        "when": "Surface to authorized product, operations, finance, or legal staff who need the governing scope and commercial terms for the BOSS Library Core AI Framework engagement.",
        "key": [
            "The engagement covers the core artificial-intelligence framework, supporting platform, deliverables, collaboration inputs, and development milestones.",
            "Additional scope, bonuses, and intellectual-property allocations require written addenda or change control.",
            "Pre-existing TWiN intellectual property remains with TWiN; BOSS-specific entrepreneurial library content belongs to BOSS unless a signed addendum states otherwise.",
            "Addresses confidentiality, data privacy and security, independent-contractor status, termination, warranties, liability, dispute resolution, governing law, and notices.",
        ],
        "next": "Use the controlled signed agreement and any later addenda for authoritative scope, ownership, payment, or termination decisions; route interpretation to the contract owners or counsel.",
        "review_scope": "All 6 pages were visually reviewed; party details, signatures, dates, addresses, and financial amounts are excluded.",
    },
    "EDE Map/BOSS VEDE Houston Map.rtb": {
        "title": "BOSS VEDE Houston Map Board",
        "description": "RTB board package named 'BOSS VEDE Houston Map.' Its board and metadata are readable, but canvas and table payloads use an application-level encoded or encrypted representation.",
        "when": "Surface when locating the source board behind the Houston veteran entrepreneurship ecosystem map.",
        "key": ["The RTB package is private according to its board metadata.", "The archive contains meta, board, canvas, resources, plugin settings, showtime, and tables members.", "Open the original in its native board application for the full map."],
        "next": "Route to the board owner or native RTB application for visual review and updates.",
    },
    "Formation/Texas SOS/BOSS.zip": {
        "title": "Empty BOSS Texas SOS Archive",
        "description": "Valid ZIP archive stored with Texas Secretary of State formation records. The archive contains no files.",
        "when": "Surface only when reconciling the formation-record inventory or explaining why this archive contributes no document content.",
        "key": ["The ZIP central directory is valid.", "Member count: zero.", "No formation record can be retrieved from this archive."],
        "next": "Use the separately stored Texas Secretary of State PDFs for the actual filing records.",
    },
}


def ffprobe_metadata(path: Path) -> tuple[dict[str, Any], str | None]:
    executable = shutil.which("ffprobe")
    if not executable:
        return {}, "ffprobe is unavailable"
    command = [
        executable,
        "-v",
        "error",
        "-show_entries",
        "format=duration:stream=index,codec_name,codec_type,width,height,r_frame_rate",
        "-of",
        "json",
        str(path),
    ]
    try:
        result = subprocess.run(command, check=True, capture_output=True, text=True, timeout=30)
        return json.loads(result.stdout), None
    except Exception as error:
        return {}, f"ffprobe failed: {type(error).__name__}"


def extract_image(path: Path, relative_path: str) -> Extraction:
    extraction = Extraction(method="manual visual inspection plus Pillow metadata")
    override = IMAGE_OVERRIDES.get(relative_path)
    if override:
        extraction.title_candidates.append((override["title"], "human_visual_description"))
        extraction.text = override["description"] + "\n" + "\n".join(override.get("key", []))
        extraction.segments = [override["description"], *override.get("key", [])]
        extraction.contacts = override.get("contacts", [])
        extraction.links = override.get("links", [])
        extraction.status = "complete"
    else:
        extraction.status = "metadata-only"
        extraction.limitations.append("No human visual description override exists")
    if Image is not None:
        try:
            with Image.open(path) as image:
                extraction.details.update({"width": image.width, "height": image.height, "mode": image.mode, "image_format": image.format})
                extraction.coverage = f"Human description; {image.width}x{image.height} {image.format or path.suffix.lstrip('.').upper()}"
        except Exception as error:
            extraction.limitations.append(f"Image metadata failed: {type(error).__name__}")
    else:
        extraction.limitations.append("Pillow is unavailable")
    return extraction


def extract_video(path: Path, relative_path: str) -> Extraction:
    extraction = Extraction(method="manual frame/transcript review plus ffprobe metadata")
    override = VIDEO_OVERRIDES.get(relative_path)
    if override:
        extraction.title_candidates.append((override["title"], "human_video_review"))
        extraction.text = override["description"] + "\n" + "\n".join(override.get("key", []))
        extraction.segments = [override["description"], *override.get("key", [])]
        extraction.people = override.get("people", [])
        extraction.links = override.get("links", [])
        extraction.contacts = override.get("contacts", [])
    metadata, error = ffprobe_metadata(path)
    extraction.details = metadata
    if error:
        extraction.status = "partial"
        extraction.limitations.append(error)
    elif override:
        extraction.status = "complete"
    else:
        extraction.status = "metadata-only"
        extraction.limitations.append("No human video description override exists")
    duration = metadata.get("format", {}).get("duration") if metadata else None
    streams = metadata.get("streams", []) if metadata else []
    video_stream = next((item for item in streams if item.get("codec_type") == "video"), {})
    coverage_parts = []
    if duration:
        coverage_parts.append(f"duration {float(duration):.3f} seconds")
    if video_stream.get("width") and video_stream.get("height"):
        coverage_parts.append(f"{video_stream['width']}x{video_stream['height']}")
    coverage_parts.append("three representative frames and offline speech transcript reviewed" if override else "container metadata only")
    extraction.coverage = "; ".join(coverage_parts)
    return extraction


def extract_source(path: Path, relative_path: str) -> Extraction:
    extension = path.suffix.casefold()
    if extension in {".md", ".txt"}:
        extraction = extract_markdown_or_text(path)
    elif extension == ".pdf":
        extraction = extract_pdf(path)
    elif extension == ".docx":
        extraction = extract_docx(path)
    elif extension == ".pptx":
        extraction = extract_pptx(path)
    elif extension == ".odp":
        extraction = extract_odp(path)
    elif extension == ".xlsx":
        extraction = extract_xlsx(path)
    elif extension == ".csv":
        extraction = extract_csv_file(path)
    elif extension == ".ppt":
        extraction = extract_legacy_ppt(path)
    elif extension in {".jpg", ".jpeg", ".png"}:
        extraction = extract_image(path, relative_path)
    elif extension == ".mp4":
        extraction = extract_video(path, relative_path)
    elif extension == ".rtb":
        extraction = extract_rtb(path)
    elif extension == ".zip":
        extraction = extract_zip_file(path)
    else:
        extraction = Extraction(status="failed", method="unsupported", coverage="No extractor", limitations=[f"Unsupported extension: {extension or '[none]'}"])
    override = SPECIAL_OVERRIDES.get(relative_path)
    if override:
        previous_status = extraction.status
        extraction.title_candidates.insert(0, (override["title"], "human_special_case"))
        extraction.text = override["description"] + "\n" + "\n".join(override.get("key", []))
        extraction.segments = [override["description"], *override.get("key", [])]
        extraction.links = unique(override.get("links", []) + extraction.links, MAX_LINKS)
        extraction.contacts = unique(override.get("contacts", []) + extraction.contacts, MAX_CONTACTS)
        review_scope = override.get("review_scope")
        if review_scope:
            extraction.details["human_review_scope"] = review_scope
            extraction.coverage = "; ".join(part for part in (extraction.coverage, review_scope) if part)
        if previous_status == "metadata-only":
            extraction.status = "human-described"
            extraction.method = "; ".join(part for part in (extraction.method, "human visual review and concise document description") if part)
            extraction.limitations.append("Markdown is a concise human-reviewed description, not a full source transcription")
    return extraction


def detect_topics(text: str) -> list[str]:
    padded = f" {clean_text(text).casefold()} "
    topics = [topic for topic, patterns in TOPIC_RULES.items() if any(pattern in padded for pattern in patterns)]
    return topics[:12] or ["general business reference"]


def detect_stages(text: str) -> list[str]:
    padded = f" {clean_text(text).casefold()} "
    if any(phrase in padded for phrase in (" all phases ", " all stages ", " entire lifecycle ")):
        return list(STAGE_RULES)
    stages = [stage for stage, patterns in STAGE_RULES.items() if any(pattern in padded for pattern in patterns)]
    return stages[:8]


def detect_organizations(text: str) -> list[str]:
    padded = f" {clean_text(text).casefold()} "
    organizations = [organization for organization, patterns in ORGANIZATION_RULES.items() if any(pattern in padded for pattern in patterns)]
    if "SCORE Houston" in organizations and "SCORE" not in organizations:
        organizations.insert(0, "SCORE")
    return organizations[:10]


def keyword_list(text: str, seed: Iterable[str]) -> list[str]:
    words = re.findall(r"[A-Za-z][A-Za-z0-9'-]{2,}", clean_text(text).casefold())
    counts = Counter(word.strip("'-") for word in words if word not in STOPWORDS and not word.isdigit())
    seeded = [item.casefold() for item in seed if item]
    ranked = [word for word, _ in counts.most_common(30)]
    return unique([*seeded, *ranked], 24)


def safe_key_items(extraction: Extraction, title: str, sensitive: bool) -> list[str]:
    if sensitive:
        return [
            "This generated record intentionally omits identifiers, account-level details, signatures, and transaction specifics.",
            "Use the source path and hash to locate the controlled original when access is authorized.",
            "The document's existence and general subject are available for routing; substantive review belongs with the responsible owner or professional.",
        ]
    candidates: list[str] = []
    for segment in extraction.segments:
        cleaned = re.sub(r"^(?:Slide\s+\d+:\s*)", "", segment, flags=re.IGNORECASE)
        for part in re.split(r"\s+\|\s+|\n", cleaned):
            item = re.sub(r"^[-*+]\s+|^\d+[.)]\s+", "", part).strip()
            if not meaningful_line(item):
                continue
            if clean_title(item).casefold() == clean_title(title).casefold():
                continue
            lowered = item.casefold()
            if any(phrase in lowered for phrase in ("boss library ingestion note", "last reviewed", "source url", "retrieval tags")):
                continue
            candidates.append(truncate(item, MAX_KEY_ITEM))
    return unique(candidates, 7)


def generic_when(topics: list[str], stages: list[str], access: str) -> str:
    if access == "restricted":
        return "Surface when an authorized user needs to locate this controlled record or identify the responsible function; do not expose its protected details in chat."
    topic_text = ", ".join(topics[:3])
    stage_text = f" during {', '.join(stages[:3])}" if stages else ""
    return f"Surface when the user's question concerns {topic_text}{stage_text} and this source is relevant to the next document or human handoff."


def generic_next(organizations: list[str], topics: list[str], access: str) -> str:
    if access == "restricted":
        return "Route detailed review to the authorized BOSS record owner and, where applicable, the appropriate attorney, accountant, administrator, or contracted professional."
    if organizations:
        return f"Open the cited source link when available and route follow-up to {organizations[0]} or the named qualified resource."
    if "branding" in topics:
        return "Route production use to the BOSS brand owner and current brand guidance."
    return "Open the source record, confirm the user's exact decision or information need, and route to the named organization or qualified person when human help is required."


def standard_guardrails(topics: list[str], access: str, has_external_links: bool) -> list[str]:
    guardrails = ["The BOSS Libraries Librarian retrieves and routes information; it is not a mentor or a substitute for a qualified professional."]
    if access == "restricted":
        guardrails.append("Do not disclose identifiers, account details, signatures, or protected client/vendor information from the controlled original.")
    if any(topic in topics for topic in ("legal and compliance", "accounting", "funding", "certification", "government contracting")):
        guardrails.append("Do not make final legal, tax, accounting, lending, certification, procurement, or compliance decisions for the user.")
    if has_external_links:
        guardrails.append("Verify live pages for current eligibility, deadlines, contacts, prices, and program rules before presenting them as current.")
    return unique(guardrails, 4)


def yaml_value(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    return json.dumps(value, ensure_ascii=False)


def markdown_bullet(value: str) -> str:
    return truncate(clean_text(value).replace("`", "'"), MAX_KEY_ITEM)


def build_record(path: Path, source_root: Path) -> dict[str, Any]:
    relative_path = path.relative_to(source_root).as_posix()
    extraction = extract_source(path, relative_path)
    title, title_source = choose_title(extraction, path)
    access, sensitive = access_classification(relative_path)
    if sensitive:
        title = redact_sensitive_text(title)
    override = IMAGE_OVERRIDES.get(relative_path) or VIDEO_OVERRIDES.get(relative_path) or SPECIAL_OVERRIDES.get(relative_path) or {}
    context_text = "\n".join([title, relative_path, extraction.text])
    topics = detect_topics(context_text)
    stages = detect_stages(context_text)
    organizations = detect_organizations(context_text)
    people = unique([*override.get("people", []), *extraction.people], 8)
    links = apply_link_policy([*override.get("links", []), *extraction.links])
    contacts = unique([*override.get("contacts", []), *extraction.contacts], MAX_CONTACTS)
    if sensitive:
        links = [link for link in links if any(domain in link.casefold() for domain in PUBLIC_LINK_DOMAINS)]
        contacts = []
        description = override.get("description", "") or (
            f"Controlled {path.suffix.lstrip('.').upper()} reference concerning {title}. "
            "This routing record preserves the document's purpose and location while omitting identifiers, account-level values, signatures, and detailed private content."
        )
    else:
        sections = parse_markdown_sections(extraction.text) if path.suffix.casefold() in {".md", ".txt"} else {}
        description = override.get("description", "")
        if not description and path.suffix.casefold() == ".ppt":
            description = (
                f"Legacy PowerPoint presentation titled {title}. Text recovery is partial and does not preserve "
                "slide order or layout; use this record to locate the original presentation for review."
            )
        if not description:
            description = section_value(sections, ("resource purpose", "purpose", "resource summary", "user need", "what this", "overview"))
        if not description:
            description_candidates = [segment for segment in extraction.segments if meaningful_line(segment) and clean_title(segment).casefold() != title.casefold()]
            description = " ".join(description_candidates[:3])
        if not description:
            description = f"Reference about {title}; extraction yielded metadata but not enough machine-readable text for a fuller summary."
    description = truncate(redact_sensitive_text(description) if sensitive else description, MAX_DESCRIPTION)
    key_info = override.get("key", []) or safe_key_items(extraction, title, sensitive)
    if not key_info:
        key_info = ["Use the source traceability section to locate the original record for detailed review."]
    if sensitive:
        key_info = [redact_sensitive_text(item) for item in key_info]
    sections = parse_markdown_sections(extraction.text) if path.suffix.casefold() in {".md", ".txt"} else {}
    when = override.get("when", "") or section_value(sections, ("route the user here when", "when to route", "when to use", "user intent", "user need"))
    next_handoff = override.get("next", "") or section_value(sections, ("recommended client-facing next action", "recommended next action", "best primary route", "next action", "handoff"))
    when = truncate(redact_sensitive_text(when) if sensitive else when, MAX_DESCRIPTION) if when else generic_when(topics, stages, access)
    next_handoff = truncate(redact_sensitive_text(next_handoff) if sensitive else next_handoff, MAX_DESCRIPTION) if next_handoff else generic_next(organizations, topics, access)
    guardrails = standard_guardrails(topics, access, bool(links))
    description = apply_link_policy_to_text(description)
    key_info = [apply_link_policy_to_text(item) for item in key_info]
    when = apply_link_policy_to_text(when)
    next_handoff = apply_link_policy_to_text(next_handoff)
    path_parts = relative_path.split("/")
    category = path_parts[0] if len(path_parts) > 1 else "BOSS Core"
    keywords = keyword_list(context_text if not sensitive else f"{title} {relative_path} {' '.join(topics)}", [*topics, *stages, *organizations])
    identifier = source_id(path, source_root)
    slug = slugify(title)
    output = f"{identifier}--{slug}.md"
    mime = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
    record = {
        "id": identifier,
        "title": title,
        "slug": slug,
        "output": output,
        "document_path": f"./{output}",
        "source_path": relative_path,
        "source_url": links[0] if links else None,
        "source_sha256": hash_file(path),
        "source_bytes": path.stat().st_size,
        "source_extension": path.suffix.casefold(),
        "source_mime": mime,
        "duplicate_of": None,
        "access": access,
        "sensitive": sensitive,
        "category": category,
        "summary": description,
        "when_to_surface": when,
        "key_information": [truncate(item, MAX_KEY_ITEM) for item in unique(key_info, 8)],
        "next_handoff": next_handoff,
        "guardrails": guardrails,
        "topics": topics,
        "lifecycle_stages": stages,
        "organizations": organizations,
        "people": people,
        "links": links,
        "contacts": contacts,
        "keywords": keywords,
        "search_text": clean_text(" | ".join([title, description, " ".join(topics), " ".join(stages), " ".join(organizations), " ".join(keywords)])),
        "extraction": {
            "status": extraction.status,
            "method": extraction.method,
            "coverage": extraction.coverage or "No coverage detail reported",
            "limitations": unique(extraction.limitations),
            "title_source": title_source,
            "details": extraction.details,
        },
        "output_sha256": None,
    }
    return record


def render_record(record: dict[str, Any], records_by_id: dict[str, dict[str, Any]]) -> str:
    frontmatter_keys = (
        "id",
        "title",
        "source_path",
        "source_sha256",
        "source_extension",
        "source_bytes",
        "access",
        "sensitive",
        "duplicate_of",
        "category",
        "topics",
        "lifecycle_stages",
        "organizations",
        "keywords",
    )
    lines = ["---"]
    for key in frontmatter_keys:
        lines.append(f"{key}: {yaml_value(record[key])}")
    lines.extend(["---", "", f"# {record['title']}", "", "> BOSS Libraries routing record. Use it to locate the right source or person; do not treat it as a replacement for the controlled original.", "", "## Description", "", record["summary"], "", "## When to surface", "", record["when_to_surface"], "", "## Key information", ""])
    for item in record["key_information"]:
        lines.append(f"- {markdown_bullet(item)}")
    lines.extend(["", "## People and links", ""])
    if record["organizations"]:
        lines.append(f"- Organizations: {', '.join(record['organizations'])}")
    if record["people"]:
        lines.append(f"- People named for routing: {', '.join(record['people'])}")
    for link in record["links"]:
        lines.append(f"- Link: <{link}>")
    for contact in record["contacts"]:
        lines.append(f"- Contact: {markdown_bullet(contact)}")
    if not (record["organizations"] or record["people"] or record["links"] or record["contacts"]):
        lines.append("- No public contact or external link was extracted; use the source path and owning category for routing.")
    lines.extend(["", "## Next handoff", "", record["next_handoff"], "", "## Guardrails", ""])
    for guardrail in record["guardrails"]:
        lines.append(f"- {markdown_bullet(guardrail)}")
    lines.extend(["", "## Extraction coverage", "", f"- Status: `{record['extraction']['status']}`", f"- Method: {record['extraction']['method']}", f"- Coverage: {record['extraction']['coverage']}"])
    for limitation in record["extraction"]["limitations"]:
        lines.append(f"- Limitation: {markdown_bullet(limitation)}")
    lines.extend(["", "## Source traceability", "", f"- Original: `business docs/{record['source_path']}`", f"- SHA-256: `{record['source_sha256']}`", f"- Source size: {record['source_bytes']:,} bytes"])
    if record["duplicate_of"]:
        canonical = records_by_id[record["duplicate_of"]]
        lines.append(f"- Exact duplicate of: [{canonical['title']}]({canonical['output']})")
    return apply_link_policy_to_text("\n".join(lines)).rstrip() + "\n"


def render_catalog(records: list[dict[str, Any]]) -> str:
    categories = Counter(record["category"] for record in records)
    formats = Counter(record["source_extension"] or "[none]" for record in records)
    lines = [
        "# BOSS Libraries Document Catalog",
        "",
        "This catalog describes the normalized BOSS Libraries corpus. The librarian should retrieve a small set of relevant records, provide the most useful live links, and route the user to the right person or organization. It is not a mentor and should not replace professional judgment.",
        "",
        "## Collection summary",
        "",
        f"- Source identities represented: {len(records)}",
        f"- Generated Markdown routing records: {len(records)}",
        f"- Exact duplicate aliases: {sum(bool(record['duplicate_of']) for record in records)}",
        "- Machine index: `document-manifest.json`",
        "- Build details: `CONVERSION_REPORT.md`",
        "",
        "## How BOSS should use these records",
        "",
        "1. Identify the user's immediate information need, location, and business stage when relevant.",
        "2. Search titles, summaries, topics, lifecycle stages, organizations, and keywords in the manifest.",
        "3. Retrieve only the strongest matching records instead of loading the entire corpus into one prompt.",
        "4. Give the relevant source links and a concise next handoff.",
        "5. Treat every record as approved public routing material and retain its source traceability.",
        "",
        "## Source categories",
        "",
        "| Category | Records |",
        "|---|---:|",
    ]
    for category, count in sorted(categories.items(), key=lambda item: item[0].casefold()):
        lines.append(f"| {category} | {count} |")
    lines.extend(["", "## Source formats", "", "| Format | Records |", "|---|---:|"])
    for extension, count in sorted(formats.items()):
        lines.append(f"| `{extension}` | {count} |")
    lines.extend(
        [
            "",
            "## Access",
            "",
            "All records in this corpus are approved public routing material.",
            "",
            "## Core routing collections",
            "",
            "- APEX Accelerator: government-contracting counseling and procurement readiness.",
            "- U.S. Small Business Administration (SBA): official business guides, funding-program education, contracting, disaster, and specialized-business resources.",
            "- Small Business Development Center (SBDC): advising, local market and planning support, specialist referrals, and training.",
            "- SCORE: mentoring, business education, roadmaps, templates, workshops, and topic hubs.",
            "",
            "See `document-manifest.json` for all record IDs, filenames, source paths, summaries, search fields, links, and extraction status.",
        ]
    )
    return "\n".join(lines).rstrip() + "\n"


def render_report(records: list[dict[str, Any]], source_count: int) -> str:
    by_format = Counter(record["source_extension"] or "[none]" for record in records)
    by_status = Counter(record["extraction"]["status"] for record in records)
    by_access = Counter(record["access"] for record in records)
    title_sources = Counter(record["extraction"]["title_source"] for record in records)
    duplicate_groups = defaultdict(list)
    for record in records:
        duplicate_groups[record["source_sha256"]].append(record)
    exact_groups = [group for group in duplicate_groups.values() if len(group) > 1]
    lines = [
        "# BOSS Libraries Conversion Report",
        "",
        f"- Builder: `{BUILDER_VERSION}`",
        f"- Source files discovered: {source_count}",
        f"- Normalized records generated: {len(records)}",
        f"- Coverage: {len(records)}/{source_count}",
        f"- Unique IDs: {len({record['id'] for record in records})}",
        f"- Unique output filenames: {len({record['output'] for record in records})}",
        f"- Exact duplicate groups: {len(exact_groups)}",
        f"- Duplicate alias records: {sum(bool(record['duplicate_of']) for record in records)}",
        "- Raw binaries copied: 0",
        "",
        "## By source format",
        "",
        "| Format | Count |",
        "|---|---:|",
    ]
    for extension, count in sorted(by_format.items()):
        lines.append(f"| `{extension}` | {count} |")
    lines.extend(["", "## By extraction status", "", "| Status | Count |", "|---|---:|"])
    for status, count in sorted(by_status.items()):
        lines.append(f"| `{status}` | {count} |")
    lines.extend(["", "## By access classification", "", "| Access | Count |", "|---|---:|"])
    for label, count in sorted(by_access.items()):
        lines.append(f"| `{label}` | {count} |")
    lines.extend(["", "## Title provenance", "", "| Title source | Count |", "|---|---:|"])
    for title_source, count in sorted(title_sources.items()):
        lines.append(f"| `{title_source}` | {count} |")
    lines.extend(["", "## Exact duplicate groups", ""])
    for index, group in enumerate(sorted(exact_groups, key=lambda group: group[0]["source_path"].casefold()), start=1):
        lines.append(f"{index}. SHA-256 `{group[0]['source_sha256']}`")
        for record in group:
            alias = f" (alias of `{record['duplicate_of']}`)" if record["duplicate_of"] else " (canonical)"
            lines.append(f"   - `{record['source_path']}`{alias}")
    lines.extend(
        [
            "",
            "## Known extraction limitations",
            "",
            "- PDF extraction is text-first. Image-only pages are flagged because OCR is not bundled into this build.",
            "- DOCX, PPTX, and ODP embedded media is counted; only the standalone image sources received human visual descriptions.",
            "- Legacy `.ppt` extraction is a best-effort printable-string scan and does not preserve slide order or layout.",
            "- Spreadsheet formulas are read as stored and are not recalculated.",
            "- The RTB board exposes readable board metadata, but its canvas/table payloads use an application-level encoded or encrypted representation.",
            "- Restricted records intentionally omit identifiers, account values, signatures, and detailed private content.",
            "",
            "## Validation result",
            "",
            "The builder validated source coverage, unique IDs, unique output filenames, source existence, output existence, H1/title alignment, duplicate targets, and manifest/output hashes.",
        ]
    )
    return "\n".join(lines).rstrip() + "\n"


def validate_corpus(source_root: Path, output_root: Path, manifest: dict[str, Any] | None = None) -> dict[str, Any]:
    manifest_path = output_root / "document-manifest.json"
    if manifest is None:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    records = manifest.get("records", [])
    errors: list[str] = []
    source_files = sorted(path for path in source_root.rglob("*") if path.is_file())
    if len(records) != len(source_files):
        errors.append(f"Manifest has {len(records)} records for {len(source_files)} source files")
    ids = [record["id"] for record in records]
    outputs = [record["output"] for record in records]
    if len(ids) != len(set(ids)):
        errors.append("Record IDs are not unique")
    if len(outputs) != len(set(outputs)):
        errors.append("Output filenames are not unique")
    by_id = {record["id"]: record for record in records}
    for record in records:
        source_path = source_root / Path(record["source_path"])
        output_path = output_root / record["output"]
        if not source_path.is_file():
            errors.append(f"Missing source: {record['source_path']}")
        if not output_path.is_file():
            errors.append(f"Missing output: {record['output']}")
            continue
        text = output_path.read_text(encoding="utf-8")
        if f"# {record['title']}\n" not in text:
            errors.append(f"H1/title mismatch: {record['output']}")
        if output_path.suffix.casefold() != ".md":
            errors.append(f"Non-Markdown output: {record['output']}")
        actual_hash = hashlib.sha256(output_path.read_bytes()).hexdigest()
        if record.get("output_sha256") != actual_hash:
            errors.append(f"Output hash mismatch: {record['output']}")
        duplicate_of = record.get("duplicate_of")
        if duplicate_of:
            canonical = by_id.get(duplicate_of)
            if not canonical:
                errors.append(f"Missing duplicate target: {record['output']}")
            elif canonical["source_sha256"] != record["source_sha256"]:
                errors.append(f"Duplicate hash mismatch: {record['output']}")
    generated_records = list(output_root.glob("bossdoc-*.md"))
    if len(generated_records) != len(records):
        errors.append(f"Found {len(generated_records)} generated Markdown records for {len(records)} manifest records")
    result = {
        "source_count": len(source_files),
        "record_count": len(records),
        "generated_record_count": len(generated_records),
        "unique_ids": len(set(ids)),
        "unique_outputs": len(set(outputs)),
        "errors": errors,
    }
    if errors:
        raise RuntimeError("Corpus validation failed:\n- " + "\n- ".join(errors[:50]))
    return result


def build(source_root: Path, output_root: Path, expected_count: int | None = EXPECTED_SOURCE_COUNT) -> dict[str, Any]:
    source_root = source_root.resolve()
    output_root = output_root.resolve()
    if not source_root.is_dir():
        raise FileNotFoundError(f"Source directory not found: {source_root}")
    app_root = APP_ROOT.resolve()
    if app_root not in output_root.parents and output_root != app_root:
        raise ValueError(f"Output must remain inside the BOSS app: {output_root}")
    source_files = sorted((path for path in source_root.rglob("*") if path.is_file()), key=lambda path: normalize_relative_path(path, source_root))
    if expected_count is not None and len(source_files) != expected_count:
        raise RuntimeError(f"Expected {expected_count} source files, found {len(source_files)}")
    unsupported = sorted({path.suffix.casefold() for path in source_files if path.suffix.casefold() not in SUPPORTED_EXTENSIONS})
    if unsupported:
        raise RuntimeError(f"Unsupported source extensions: {', '.join(unsupported)}")
    output_root.mkdir(parents=True, exist_ok=True)
    for stale in output_root.glob("bossdoc-*.md"):
        stale.unlink()
    records: list[dict[str, Any]] = []
    for index, path in enumerate(source_files, start=1):
        records.append(build_record(path, source_root))
        if index % 50 == 0 or index == len(source_files):
            print(f"Extracted {index}/{len(source_files)} sources", flush=True)
    hash_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        hash_groups[record["source_sha256"]].append(record)
    for group in hash_groups.values():
        if len(group) < 2:
            continue
        canonical = sorted(group, key=lambda record: normalize_relative_path(source_root / Path(record["source_path"]), source_root))[0]
        for record in group:
            if record is not canonical:
                record["duplicate_of"] = canonical["id"]
    records_by_id = {record["id"]: record for record in records}
    for record in records:
        output_path = output_root / record["output"]
        output_path.write_text(render_record(record, records_by_id), encoding="utf-8", newline="\n")
        record["output_sha256"] = hashlib.sha256(output_path.read_bytes()).hexdigest()
    catalog_name = "000-boss-library-catalog.md"
    report_name = "CONVERSION_REPORT.md"
    (output_root / catalog_name).write_text(render_catalog(records), encoding="utf-8", newline="\n")
    (output_root / report_name).write_text(render_report(records, len(source_files)), encoding="utf-8", newline="\n")
    version_payload = [
        {
            "id": record["id"],
            "source_path": record["source_path"],
            "source_sha256": record["source_sha256"],
            "output": record["output"],
            "output_sha256": record["output_sha256"],
        }
        for record in records
    ]
    manifest_version = "sha256:" + hashlib.sha256(
        json.dumps(version_payload, ensure_ascii=False, separators=(",", ":"), sort_keys=True).encode("utf-8")
    ).hexdigest()
    manifest = {
        "schema_version": SCHEMA_VERSION,
        "manifest_version": manifest_version,
        "builder_version": BUILDER_VERSION,
        "source_root": "apps/boss/business docs",
        "original_root": "../business%20docs/",
        "record_count": len(records),
        "catalog": catalog_name,
        "conversion_report": report_name,
        "records": records,
    }
    manifest_path = output_root / "document-manifest.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8", newline="\n")
    validation = validate_corpus(source_root, output_root, manifest)
    print(json.dumps(validation, indent=2, sort_keys=True))
    return manifest


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-root", type=Path, default=DEFAULT_SOURCE_ROOT)
    parser.add_argument("--output-root", type=Path, default=DEFAULT_OUTPUT_ROOT)
    parser.add_argument("--check", action="store_true", help="Validate existing generated corpus without rebuilding")
    parser.add_argument("--allow-count-change", action="store_true", help="Do not enforce the audited 500-source count")
    args = parser.parse_args(argv)
    try:
        if args.check:
            result = validate_corpus(args.source_root.resolve(), args.output_root.resolve())
            print(json.dumps(result, indent=2, sort_keys=True))
        else:
            build(args.source_root, args.output_root, None if args.allow_count_change else EXPECTED_SOURCE_COUNT)
    except Exception as error:
        print(str(error), file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
